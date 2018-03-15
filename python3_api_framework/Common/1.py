from openpyxl import load_workbook
import os
class DoExcel:

    def __init__(self,excelpath):
        wb = load_workbook(excelpath)
        self.sh_case_data  = wb["case_data"]
        self.init_data = wb["init_data"]


    def sl(self):

         all_init_dada = {}
         print(type(self.init_data))
         for index in range(2, self.init_data.max_row + 1):
             print(index)
             print(type(index))

    def get_init_data(self):
        all_init_dada = {}
        for index in range(2, self.init_data.max_row + 1):
            key = self.init_data.cell(row=index, column=1).value
            all_init_dada[key] = str(self.init_data.cell(row=index, column=2).value)

        all_init_dada['$phone'] = str(int(all_init_dada['$init_phone']) + 1)
        return all_init_dada

    def get_caseData_all(self):
        all_case_datas = []
        for index in range(2, self.sh_case_data.max_row + 1):
            print('分行读取数据',index)
            case_data = {}
            case_data['id'] = self.sh_case_data.cell(row=index, column=1).value
            case_data['method'] = self.sh_case_data.cell(row=index, column=5).value
            case_data['url'] = self.sh_case_data.cell(row=index, column=6).value
            temp_cese_data = self.sh_case_data.cell(row=index, column=7).value
            init_data = self.get_init_data()
            print('初始化数据', init_data)
            for key, value in init_data.items():
                print(key,value)
                if temp_cese_data.find(key) != -1:
                    temp_cese_data = temp_cese_data.replace(key, value)
            case_data['request_data'] = temp_cese_data
            case_data['response_data'] = self.sh_case_data.cell(row=index, column=8).value
            all_case_datas.append(case_data)
        return all_case_datas


de = DoExcel(os.getcwd().replace('Common','TestDatas')+'/1.xlsx')
h = de.get_caseData_all()
print(h)
