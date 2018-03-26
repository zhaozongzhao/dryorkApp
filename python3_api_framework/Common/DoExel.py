from openpyxl import load_workbook
import os
class DoExcel:

    def __init__(self,excelpath,logger):
        #初始化数据，打开excel表格
        self.wb = load_workbook(excelpath)
        #打开相关sheet页
        self.sh_case_data  = self.wb["case_data"]
        self.init_data =self.wb["init_data"]
        self.log = logger

    def get_init_data(self):
        #查询配置数据用来替换case数据中的标识
        all_init_dada = {}
        for index in range(2, self.init_data.max_row + 1):
            key = self.init_data.cell(row=index, column=1).value
            all_init_dada[key] = str(self.init_data.cell(row=index, column=2).value)
        #对配置数据进行整理
        all_init_dada['$phone'] = str(int(all_init_dada['$init_phone']) + 1)
        all_init_dada['$phone1'] = str(int(all_init_dada['$init_phone']) + 2)
        self.log.info('配置数据'+str(all_init_dada))
        return all_init_dada

    def get_caseData_all(self):
        #查询casedata数据，并将数据以字典的方式返回
        all_case_datas = []
        for index in range(2, self.sh_case_data.max_row + 1):
            #读取数据从第二行开始到最大行数加1
            self.log.info('分行读取数据'+str(index))
            case_data = {}
            case_data['id'] = self.sh_case_data.cell(row=index, column=1).value
            case_data['api_name'] = self.sh_case_data.cell(row=index, column=4).value
            case_data['method'] = self.sh_case_data.cell(row=index, column=5).value
            case_data['url'] = self.sh_case_data.cell(row=index, column=6).value
            temp_cese_data = self.sh_case_data.cell(row=index, column=7).value
            init_data = self.get_init_data()
            self.log.info('初始化数据'+str(init_data))
            #判断读取的测试数据中是否存在标识，存在替换
            if temp_cese_data is not None:
               for key, value in init_data.items():
                  print(key, value)
                  if temp_cese_data.find(key) != -1:
                    #使用find函数查询字符串中是否包含子字符串
                    temp_cese_data = temp_cese_data.replace(key, value)
                    self.log.info('替换后的测试数据'+str(temp_cese_data))
            case_data['request_data'] = temp_cese_data
            case_data['response_data'] = self.sh_case_data.cell(row=index, column=8).value
            case_data['is_all'] = self.sh_case_data.cell(row=index, column=9).value
            if self.sh_case_data.cell(row=index, column=10).value is not None:
                case_data['related_exp']=self.sh_case_data.cell(row=index, column=10)
            all_case_datas.append(case_data)
            self.log.info('测试数据'+str(all_case_datas))
        return all_case_datas

    def Modify_init_data(self):
        #修改初始数据
         self.log.info('测试结束修改初始数据')
         init_data = self.get_init_data()
         value =str(int(init_data['$init_phone'])+3)
         self.init_data.cell(row=2, column=2).value = value

    def save_excel(self,excelpath):
         #保存表格
         #sava保存文件并且关闭文件
         self.log.info('测试结束关闭文件')
         self.wb.save(excelpath)


if __name__ == '__main__':

  de = DoExcel(os.path.split(os.path.realpath(__file__))[0].replace('Common', 'TestDatas') + 'api_qcd.xlsx')
  all_case_datas = de.get_caseData_all()
  print(all_case_datas)
  de.Modify_init_data()
  de.save_excel(os.path.split(os.path.realpath(__file__))[0].replace('Common', 'TestDatas') + 'api_qcd.xlsx')